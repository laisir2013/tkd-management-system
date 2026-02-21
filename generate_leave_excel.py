import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

# Load data
with open('elite_leaves.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Merge similar names (same person with different capitalization/spacing)
# Group all dates by year
year_data = {2024: defaultdict(list), 2025: defaultdict(list), 2026: defaultdict(list)}

for name, dates in data.items():
    for d in dates:
        year = int(d[:4])
        if year in year_data:
            year_data[year][name].append(d)

# Sort dates within each name
for year in year_data:
    for name in year_data[year]:
        year_data[year][name].sort()

# Create workbook
wb = Workbook()

# Style definitions
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
cell_align = Alignment(horizontal="center", vertical="center")
name_font = Font(bold=True, size=11)
date_font = Font(size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for idx, year in enumerate([2024, 2025, 2026]):
    if idx == 0:
        ws = wb.active
        ws.title = str(year)
    else:
        ws = wb.create_sheet(title=str(year))
    
    students = year_data[year]
    if not students:
        ws.cell(row=1, column=1, value=f"{year}年 無請假記錄")
        continue
    
    # Find max number of leave dates for any student this year
    max_dates = max(len(dates) for dates in students.values())
    
    # Sort students: by leave count (descending), then by name
    sorted_students = sorted(students.items(), key=lambda x: (-len(x[1]), x[0]))
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_dates + 2)
    title_cell = ws.cell(row=1, column=1, value=f"創武精英班請假記錄 — {year}年 (共 {len(sorted_students)} 位學生，{sum(len(d) for d in students.values())} 筆請假)")
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Header row
    row = 2
    headers = ["姓名", "請假次數"] + [f"請假日期{i+1}" for i in range(max_dates)]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 25
    
    # Data rows
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i, (name, dates) in enumerate(sorted_students):
        row = i + 3
        fill = even_fill if i % 2 == 0 else odd_fill
        
        # Name
        cell = ws.cell(row=row, column=1, value=name)
        cell.font = name_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        cell.fill = fill
        
        # Leave count
        cell = ws.cell(row=row, column=2, value=len(dates))
        cell.font = Font(bold=True, size=11, color="CC0000")
        cell.alignment = cell_align
        cell.border = thin_border
        cell.fill = fill
        
        # Dates - format as M月D日
        for j, d in enumerate(dates):
            month = int(d[5:7])
            day = int(d[8:10])
            date_str = f"{month}月{day}日"
            cell = ws.cell(row=row, column=j + 3, value=date_str)
            cell.font = date_font
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = fill
    
    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    for col_idx in range(3, max_dates + 3):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    
    # Freeze panes (freeze header)
    ws.freeze_panes = "C3"

# Save
output_path = '/home/user/webapp/精英班請假記錄_2024_2025_2026.xlsx'
wb.save(output_path)
print(f"Excel saved: {output_path}")

# Print summary
for year in [2024, 2025, 2026]:
    students = year_data[year]
    total_leaves = sum(len(d) for d in students.values())
    print(f"\n{year}年: {len(students)} 位學生, {total_leaves} 筆請假")
    sorted_s = sorted(students.items(), key=lambda x: -len(x[1]))[:5]
    for name, dates in sorted_s:
        print(f"  {name}: {len(dates)} 次")
